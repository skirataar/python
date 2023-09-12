from openpyxl import Workbook 
from openpyxl.styles import Font

# Writing Data to workbook
wb = Workbook() # Create a Workbook
sheet = wb.active # active the created workbook
sheet.title = "Language" # set name for created sheet
wb.create_sheet(title = "Capital") # Create sheet and set name 


lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code =['KA', 'TS', 'TN']


#header of the Excel
sheet.cell(row = 1, column = 1).value = "State"
sheet.cell(row = 1, column = 2).value = "Language"
sheet.cell(row = 1, column = 3).value = "Code"


# make bold header 
ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft
        
for i in range(2,5):
    sheet.cell(row = i, column = 1).value = state[i-2]
    sheet.cell(row = i, column = 2).value = lang[i-2]
    sheet.cell(row = i, column = 3).value = code[i-2]


wb.save("C:\\Users\\deepa\\OneDrive\\Desktop\\demo.xlsx")




srchCode = input("Enter state code for finding language ")
for i in range(2,5):
    data = sheet.cell(row = i, column = 3).value
    if data == srchCode:
        print("Corresponding language for code", srchCode, "is", sheet.cell(row = i, column = 2).value)


wb.close()
